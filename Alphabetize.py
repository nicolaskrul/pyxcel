#python 3.4
#Dit scriptje haalt informatie op uit een gegeven excelbestand. 
#De velden zullen op alfabetische volgorde worden geordent, op basis van de informatie uit de eerste kolom.
#Er is geen limiet geplaatst op de hoeveelheid informatie. In principe moet het mogelijk zijn om grotere documenten te verwerken.
#
# Ontworpen door Nicolas Krul

import sys
import os.path
import openpyxl as op

def main():
  try:
      filename = str(sys.argv[1])                		#filename opvragen via commandline en doorgaan wanneer bestand gevonden is.
      wb = op.load_workbook(filename)           		#Excelbestand openen voor gebruik in openpyxl
      
      current_sheet = wb.active                  		#Eerste sheet kiezen
      new_book = op.Workbook()                   		#Nieuw excel workbook declareren

      extraction = info_extract(current_sheet)   		#Data uit excelsheet te halen
      sorted_results = sort_data(extraction)   	 		#Data aan de hand van kolom A alfabetisch sorteren 
      save_xlsx(sorted_results, new_book)        		#Opslaan in nieuw .XLSX bestand

  except IndexError:                             		#Foutmelding wanneer bestand niet bestaat.
      exit("Voer de naam van een excelbestand in: " + sys.argv[0] + ' filename.xlsx')

def info_extract(old_sheet):                     		#Functie voor het ophalen van informatie uit sheet.
   list_of_rows = []                             		#variabele voor een list bestaande uit andere lists.
   for target_row in old_sheet.iter_rows():      		#Rijen 1 voor 1 afwerken
      rows = []                                 		#list maken om rijen in op te slaan
      for column in target_row:                  		#kolom uit rij selecteren
         rows.append(column.value)               		#waarde aan "rows" lijst toevoegen
      list_of_rows.append(rows)                  		#"rows" list aan "list_of_rows" toevoegen
   return list_of_rows                           		#resultaat terug sturen voor verdere verwerking

def sort_data(data_for_processing):              		#functie voor sorteren van ruwe verzameling geextraheerde lijsten
   sorted_data = sorted(data_for_processing, key=lambda x: x[0]) #Ordenen op basis van 1e kolom. Mogelijk gemaakt door lamda functie.
   return(sorted_data)                           		#Gesorteerde data terugsturen.

def save_xlsx(data, book):                       		#Gegevens gereed maken voor export
   new_active_sheet = book.active                		#Primaire sheet van nieuw workbook gereed maken.
   for sorted_row in data:                       		#gesorteerde data in nieuwe sheet implementeren
      new_active_sheet.append(sorted_row)        
   book.save('nieuw.xlsx')                       		#Data opslaan in nieuw.xlsx
       
if __name__ == "__main__":                       		#Functie main() aanroepen wanneer script wordt gestart
    main()
